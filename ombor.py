import datetime
from openpyxl import Workbook, load_workbook
from random import randint
import random

text = """
1 - Mahsulot qo'shish.
2 - Mahsulotni ko'rish.
3 - Mahsulotni sotish.
4 - Mahsulot mudati.
5 - Xisobot.
0 - Chiqish.
"""



while True:
    wb = Workbook()
    wl = load_workbook("Baza.xlsx")
    wk = wl.active
    ws = wb.active
    Data = []
    print(text)
    sorov = int(input("Bo'lim tanlang\n>>>"))

    if sorov == 1:
        for i in range(1, wk.max_row + 1):
            A = f"A{i}"
            B = f"B{i}"
            C = f"C{i}"
            D = f"D{i}"
            E = f"E{i}"
            F = f"F{i}"
            G = f"G{i}"
            Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
        while True:
            nomi = input('Nomi: ')
            if nomi == '0':
                break
            soni = int(input('Soni: '))
            narx = int(input('Maxsulot narxi: '))
            foiz = (narx // 100) * 120
            code_sh = int(input('Shtrix kodi: '))
            now = datetime.datetime.now().strftime('%Y/%m/%d')
            a = int(input("yil:"))
            b = int(input("oy:"))
            c = int(input("kun:"))
            mudat = datetime.datetime(a, b, c).strftime("%Y/%m/%d")
            Data.append([nomi, soni, narx, foiz, code_sh, now, mudat])
        for k in Data:
            ws.append(k)
        wb.save("Baza.xlsx")


    elif sorov == 2:
        for i in range(1, wk.max_row + 1):
            A = f"A{i}"
            B = f"B{i}"
            C = f"C{i}"
            D = f"D{i}"
            E = f"E{i}"
            F = f"F{i}"
            G = f"G{i}"
            Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
        for i in Data:
            print(i)

    if sorov == 3:

        data = []
        data1 = []
        n = -1
        mahsulot_nomi = input("Mahsulot nomi: ")
        for i in range(1, wk.max_row + 1):
            A = f"A{i}"
            data.append(wk[A].value)
        for j in data:
            if mahsulot_nomi == j:
                n = data.index(j)
        if n == -1:
            print("Bizda bunday ma'lumot yoq!!!")
            continue
        for i in range(1, wk.max_row + 1):
            B = f"B{i}"
            data1.append(wk[B].value)
        sotish_soni = int(input("Sotiladigan mahsulot soni: "))
        if sotish_soni <= data1[n]:
            data1[n] -= sotish_soni
            wk.cell(row=n + 1, column=2).value = data1[n]
            for i in range(1, wk.max_row + 1):
                A = f"A{i}"
                B = f"B{i}"
                C = f"C{i}"
                D = f"D{i}"
                E = f"E{i}"
                F = f"F{i}"
                G = f"G{i}"
                Data.append([wk[A].value, wk[B].value, wk[C].value, wk[D].value, wk[E].value, wk[F].value, wk[G].value])
            for k in Data:
                ws.append(k)
            wb.save("Baza.xlsx")

        else:
            print("Bizda buncha mahsulot yo'q")


    elif sorov == 4:
        sana_1 = []
        for i in range(2, wk.max_row + 1):
            G = f"G{i}"
            sana_1.append(wk[G].value)
        for i in sana_1:
            dd = i.split('/')
            x = datetime.datetime(int(dd[0]),int(dd[1]),int(dd[2]))
            print((x-datetime.datetime.now()).days)

    elif sorov == 5:


        wb2 = load_workbook('Baza.xlsx')
        wb2.create_sheet(f"{datetime.datetime.now().strftime('%d.%m.%Y')}")
        sheet2 = wb2[f"{datetime.datetime.now().strftime('%d.%m.%Y')}"]

        sheet2['A1']='nomi'
        sheet2['B1'] = 'jami_soni'
        sheet2['C1'] = 'jami_sotilgan_narxi'
        sheet2['D1'] = 'umumiy_savdo'
        sheet2['E1'] = 'sof_foyda'


        for i in range(2,sheet2.max_row + 1):
            A = f"A{i}"
            B = f"B{i}"
            C = f"C{i}"
            D = f"D{i}"
            E = f"E{i}"




        



        wb2.save('Baza.xlsx')



    elif sorov == 0:
        print("Xayr salomat boling")
        break
